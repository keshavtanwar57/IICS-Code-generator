import streamlit as st
import pandas as pd
import openpyxl as xl 



## Logic 

# null check
def Null_Exp_Genrator():
    null_check_base_exp = "ISNULL({}),Abort('NULL CHECK FAILED {}'),"
    null_check_exp_res = r"Decode(true,"
    for ind, val in enumerate(df['Mandatory']):
        if val == 'Y':
            src = str(df['Source Field'][ind]).replace(' ','_')
            inter_res = null_check_base_exp.format(src,src)
            if inter_res != r"ISNULL(nan),Abort('NULL CHECK FAILED nan'),":
                null_check_exp_res += (inter_res + '\n')
           # print(val,df['Source Field'][ind])
          #  print(type(df['Source Field'][ind]))
    null_check_exp_res += "'0')"
    return null_check_exp_res


# len check
def Len_Check_Generator():
    len_check_base_exp = "to_integer(LENGTH({}))>TO_INTEGER('{}'),Abort('LENGTH CHECK FAILED {}'),"
    len_check_exp_res = r"Decode(true,"
    for ind,val in enumerate(df['Source Field']):
        val = str(val).replace(' ','_')
        inter_res = len_check_base_exp.format(val,df['Target Postgres Column Data Length'][ind],val)
        if "Abort('LENGTH CHECK FAILED nan')," not in inter_res and "TO_INTEGER('-')" not in inter_res and "TO_INTEGER('nan')" not in inter_res:
            len_check_exp_res += (inter_res + '\n')
    len_check_exp_res += "'0')"
    return len_check_exp_res


# UI
st.warning('Note: Make sure the file to be selected is not opened in the system')
uploaded_file = st.file_uploader("Choose a file")
header_number = st.number_input('Enter Header Number')

try:
    wb = xl.load_workbook(uploaded_file)  
    #st.write(list(wb.sheetnames))
    selected_option = st.selectbox(
    'Select Sheet',
    list(wb.sheetnames))

    df = pd.read_excel(uploaded_file,sheet_name=selected_option,header=int(header_number))
     
    st.dataframe(df, width=1500)
    if st.button('Generate Code: '):
        st.subheader('Code for Length Check: ')
        st.code(Len_Check_Generator())
        st.subheader('Code for Null Check: ')
        st.code(Null_Exp_Genrator())

except:
    st.subheader('Failed to load file. Try rechecking the file and sheet name')